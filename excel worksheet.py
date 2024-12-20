import openpyxl
import pandas as pd
# Load the workbook
df=pd.read_csv('/content/sample_data/california_housing_test.csv').head(5)
(rows,col)=df.shape
df.to_excel('/content/sample_data/california_housing_test.xlsx')
workbook = openpyxl.load_workbook('/content/sample_data/california_housing_test.xlsx')
 
# Select the worksheet
sheet = workbook['Sheet1']  # Replace 'Sheet1' with your sheet name
# sumpos='B$'+str(rows+2)
sumpos= str(rows+2)
# Write the formula to a specific cell
sheet['A'+str(rows+3)] = 'Grand Total'
sheet['B'+str(rows+3)] = f'=SUM(B$2: B${sumpos})'  # Example formula: sum of cells B1 to B10
sheet['C'+str(rows+3)] = f'=SUM(C$2: C${sumpos})'
sheet['D'+str(rows+3)] = f'=SUM(D$2: D${sumpos})'
from openpyxl.styles import Font
d4 = sheet['A'+str(rows+3)]
ft = Font(bold=True)
d4.font = ft
###############
sheet['K'+str(1)] = 'Total Long'
sheet['L'+str(1)] = '%age changes'
sheet['M'+str(1)] = 'Total Lat'
sheet['N'+str(1)] = '%age Lat changes'
for row in range(2, rows + 2):
    sheet.cell(row=row, column=col+2).value = f"=SUM(B$2: B${sumpos})"
    sheet.cell(row=row, column=col+3).value = f"=(B{row}*100)/SUM(B$2: B${sumpos})"
    sheet.cell(row=row, column=col+4).value = f"=SUM(C$2: C${sumpos})"
    sheet.cell(row=row, column=col+5).value = f"=(C{row}*100)/SUM(C$2: C${sumpos})"
##############
for cell in sheet[1]:
    cell.font=Font(bold=True,color='00808080',size=12)
    cell.style
# Save the changes
workbook.save('your_excel_file.xlsx')
